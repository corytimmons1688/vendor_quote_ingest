"""
Runs OCR on downloaded vendor quote files.
Handles PDFs, images, HTML, and Excel files.
Outputs structured JSON per file for database loading.
"""

import argparse
import json
import os
import subprocess
import sys
from pathlib import Path

import pytesseract
from PIL import Image
from pdf2image import convert_from_path
from bs4 import BeautifulSoup
import openpyxl

from vendor_extractors import extract_for_vendor


def get_tesseract_version():
    """Get the installed Tesseract version string."""
    try:
        result = subprocess.run(
            ['tesseract', '--version'],
            capture_output=True, text=True
        )
        return result.stdout.split('\n')[0]
    except Exception:
        return 'unknown'


def ocr_image(image):
    """Run OCR on a PIL Image and return text + per-word confidence data."""
    data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT)

    full_text_parts = []
    fields = []
    line_num = 0
    current_line = []

    for i in range(len(data['text'])):
        word = data['text'][i].strip()
        conf = data['conf'][i]

        if word:
            current_line.append({
                'word': word,
                'confidence': float(conf) / 100.0 if conf != -1 else None
            })

        if i + 1 >= len(data['text']) or data['line_num'][i + 1] != data['line_num'][i]:
            if current_line:
                line_num += 1
                line_text = ' '.join(w['word'] for w in current_line)
                avg_conf = sum(
                    w['confidence'] for w in current_line if w['confidence'] is not None
                ) / max(len([w for w in current_line if w['confidence'] is not None]), 1)

                full_text_parts.append(line_text)
                fields.append({
                    'line_num': line_num,
                    'text': line_text,
                    'confidence': round(avg_conf, 4),
                    'word_count': len(current_line)
                })
                current_line = []

    return '\n'.join(full_text_parts), fields


def process_pdf(filepath):
    """Convert PDF to images and OCR each page."""
    pages = convert_from_path(filepath, dpi=300)
    results = []

    for page_num, page_image in enumerate(pages, start=1):
        raw_text, fields = ocr_image(page_image)
        results.append({
            'page': page_num,
            'raw_text': raw_text,
            'fields': fields
        })

    return results


def process_image(filepath):
    """OCR a single image file."""
    image = Image.open(filepath)
    raw_text, fields = ocr_image(image)
    return [{'page': 1, 'raw_text': raw_text, 'fields': fields}]


def process_html(filepath):
    """Extract text content from HTML email body."""
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        soup = BeautifulSoup(f.read(), 'html.parser')

    text = soup.get_text(separator='\n', strip=True)
    lines = [line for line in text.split('\n') if line.strip()]

    fields = []
    for i, line in enumerate(lines, start=1):
        fields.append({
            'line_num': i,
            'text': line,
            'confidence': 1.0,
            'word_count': len(line.split())
        })

    return [{'page': 1, 'raw_text': '\n'.join(lines), 'fields': fields}]


def process_excel(filepath):
    """Extract raw cell data from Excel files."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    results = []

    for sheet_num, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        lines = []
        fields = []
        line_num = 0

        for row in ws.iter_rows(values_only=False):
            cell_values = []
            for cell in row:
                if cell.value is not None:
                    cell_values.append(str(cell.value))

            if cell_values:
                line_num += 1
                line_text = '\t'.join(cell_values)
                lines.append(line_text)
                fields.append({
                    'line_num': line_num,
                    'text': line_text,
                    'confidence': 1.0,
                    'word_count': len(cell_values)
                })

        results.append({
            'page': sheet_num,
            'sheet_name': sheet_name,
            'raw_text': '\n'.join(lines),
            'fields': fields
        })

    return results


def process_json_specs(filepath):
    """Pass through structured specification JSON files from the Apps Script."""
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)

    specs = data.get('specifications', {})
    if not specs:
        return None

    fields = []
    for i, (key, value) in enumerate(specs.items(), start=1):
        fields.append({
            'line_num': i,
            'text': f'{key}: {value}',
            'confidence': 1.0,
            'word_count': len(value.split())
        })

    raw_text = '\n'.join(f'{k}: {v}' for k, v in specs.items())
    return [{
        'page': 1,
        'raw_text': raw_text,
        'fields': fields,
        'spec_type': data.get('specType'),  # 'requested' or None
        'specifications': specs,
        'metadata': {
            'email_date': data.get('emailDate'),
            'email_subject': data.get('emailSubject'),
            'email_from': data.get('emailFrom'),
        }
    }]


# Known spec field names and their normalized column suffixes
SPEC_FIELD_MAP = {
    'bag': 'bag',
    'size': 'size',
    'substrate': 'substrate',
    'finish': 'finish',
    'material': 'material',
    'embellishment': 'embellishment',
    'fill style': 'fill_style',
    'seal type': 'seal_type',
    'gusset style': 'gusset_style',
    'gusset details': 'gusset_details',
    'zipper': 'zipper',
    'tear notch': 'tear_notch',
    'hole punch': 'hole_punch',
    'corners': 'corners',
    'printing method': 'printing_method',
    'quantities': 'quantities',
}


def extract_returned_specs(fields):
    """Scan OCR field lines for known spec field names and return matched values."""
    specs = {}
    for field in fields:
        text = field.get('text', '')
        # Try to match "Key: Value" or "Key- Value" pattern against known spec fields
        for spec_name, col_suffix in SPEC_FIELD_MAP.items():
            # Require exact field match — avoid "finishing" matching "finish"
            lower = text.lower()
            if lower.startswith(spec_name) and (
                len(lower) == len(spec_name) or
                lower[len(spec_name)] in (':', '-', '–', '—', ' ')
            ) and not lower.startswith(spec_name + 'ing'):
                # Check for colon or dash separator
                for sep in [':', '-', '–', '—']:
                    idx = text.find(sep, len(spec_name) - 1)
                    if idx > 0 and idx < len(text) - 1:
                        value = text[idx + 1:].strip()
                        if value:
                            # Split "Other: <corner spec>" out of zipper values
                            # e.g. "Presto CR Zipper Other: Round Corners"
                            if col_suffix == 'zipper' and ' Other:' in value:
                                parts = value.split(' Other:', 1)
                                specs[col_suffix] = parts[0].strip()
                                corner_val = parts[1].strip()
                                if corner_val:
                                    specs['corners'] = corner_val
                            else:
                                specs[col_suffix] = value
                            break
                break
    return specs


def process_file(filepath):
    """Route file to appropriate processor based on extension."""
    ext = Path(filepath).suffix.lower()
    processors = {
        '.pdf': process_pdf,
        '.png': process_image,
        '.jpg': process_image,
        '.jpeg': process_image,
        '.tiff': process_image,
        '.tif': process_image,
        '.html': process_html,
        '.json': process_json_specs,
        '.xlsx': process_excel,
        '.xls': process_excel,
    }

    processor = processors.get(ext)
    if processor is None:
        print(f"  WARNING: No processor for {ext}, skipping {filepath}")
        return None

    return processor(filepath)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--vendor', required=True)
    parser.add_argument('--input-dir', required=True)
    parser.add_argument('--output-dir', required=True)
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    tesseract_version = get_tesseract_version()
    input_dir = Path(args.input_dir)

    files = [
        f for f in input_dir.iterdir()
        if f.is_file() and not f.name.startswith('_') and not f.name.startswith('.')
    ]

    print(f"Processing {len(files)} files for {args.vendor}")
    summary = {'vendor': args.vendor, 'files': [], 'errors': []}

    for filepath in files:
        print(f"  Processing: {filepath.name}")
        try:
            pages = process_file(filepath)
            if pages is None:
                continue

            # Check for spec_type from JSON specs files
            spec_type = None
            specifications = None
            returned_specs = None
            for page_data in pages:
                if page_data.get('spec_type'):
                    spec_type = page_data['spec_type']
                    specifications = page_data.get('specifications')
                    break

            # For OCR'd PDFs, attempt to extract returned specs from field lines
            ext = filepath.suffix.lower()
            if ext == '.pdf' and not spec_type:
                all_fields = []
                for page_data in pages:
                    all_fields.extend(page_data.get('fields', []))
                returned_specs = extract_returned_specs(all_fields) or None

            # --- Vendor-specific structured extraction ---
            # Combine all page text for vendor extractors
            all_text = '\n'.join(p.get('raw_text', '') for p in pages)
            vendor_extracted = {}
            if all_text.strip() and not spec_type:
                vendor_extracted = extract_for_vendor(args.vendor, all_text)

            output = {
                'vendor': args.vendor,
                'source_file': filepath.name,
                'file_type': filepath.suffix.lstrip('.'),
                'file_size_bytes': filepath.stat().st_size,
                'ocr_engine': 'tesseract',
                'ocr_version': tesseract_version,
                'pages': pages
            }

            if spec_type:
                output['spec_type'] = spec_type
            if specifications:
                output['specifications'] = specifications
            if returned_specs:
                output['returned_specs'] = returned_specs
            if vendor_extracted:
                output['vendor_extracted'] = vendor_extracted

            output_path = Path(args.output_dir) / f"{filepath.stem}.json"
            with open(output_path, 'w') as f:
                json.dump(output, f, indent=2)

            summary['files'].append({
                'name': filepath.name,
                'pages': len(pages),
                'status': 'success'
            })

        except Exception as e:
            print(f"  ERROR processing {filepath.name}: {e}", file=sys.stderr)
            summary['errors'].append({
                'name': filepath.name,
                'error': str(e)
            })

    summary_path = Path(args.output_dir) / '_summary.json'
    with open(summary_path, 'w') as f:
        json.dump(summary, f, indent=2)

    print(f"Done: {len(summary['files'])} succeeded, {len(summary['errors'])} errors")


if __name__ == '__main__':
    main()
